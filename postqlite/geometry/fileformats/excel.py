


class Excel97(object):
    def __init__(self, filepath, **kwargs):
        self.filepath = filepath
        self.kwargs = kwargs

        self.reader = self.load_reader()
        self.fieldnames = self.load_fields()
        self.fieldtypes = None
        self.meta = self.load_meta()

    def __len__(self):
        if 'last' in self.kwargs:
            return self.kwargs['last']
        else:
            return self.reader.nrows
        
    def __iter__(self):
        return self.stream()

    def stream(self):
        rows = ([cell.value for cell in row] for row in self.reader.get_rows())
        
        # some excel files may contain junk metadata near top and bottom rows that should be skipped
        # TODO: maybe change API/keywords here...
        
        if "skip" in self.kwargs:
            for _ in range(self.kwargs["skip"]):
                next(rows)

        _fields = next(rows) # skip fieldnames

        if "last" in self.kwargs:
            last = self.kwargs["last"]
            rows = (r for i,r in enumerate(rows) if i <= last)

        # return along with empty geometries
        rowgeoms = ((row,None) for row in rows)
        return rowgeoms

    def load_reader(self):
        import xlrd

        encoding = self.kwargs.get('encoding', 'utf8')
        wb = xlrd.open_workbook(self.filepath, encoding_override=encoding, on_demand=True)
        
        if "sheet" in self.kwargs:
            sheet = wb.sheet_by_name(self.kwargs["sheet"])
        else:
            sheet = wb.sheet_by_index(0)
            
        return sheet

    def load_fields(self):
        rows = ([cell.value for cell in row] for row in self.reader.get_rows())

        # some excel files may contain junk metadata near top that should be skipped
        # TODO: maybe change API/keywords here...
        if "skip" in self.kwargs:
            for _ in range(self.kwargs["skip"]):
                next(rows)
                
        fieldnames = next(rows)
        return fieldnames

    def load_meta(self):
        return None


class Excel(object):
    def __init__(self, filepath, **kwargs):
        self.filepath = filepath
        self.kwargs = kwargs

        self.reader = self.load_reader()
        self.fieldnames = self.load_fields()
        self.fieldtypes = None
        self.meta = self.load_meta()

    def __len__(self):
        if 'last' in self.kwargs:
            return self.kwargs['last']
        else:
            return self.reader.nrows
        
    def __iter__(self):
        return self.stream()

    def stream(self):
        rows = ([cell.value for cell in row] for row in self.reader.iter_rows())
        
        # some excel files may contain junk metadata near top and bottom rows that should be skipped
        # TODO: maybe change API/keywords here...
        
        if "skip" in self.kwargs:
            for _ in range(self.kwargs["skip"]):
                next(rows)

        _fields = next(rows) # skip fieldnames

        if "last" in self.kwargs:
            last = self.kwargs["last"]
            rows = (r for i,r in enumerate(rows) if i <= last)

        # return along with empty geometries
        rowgeoms = ((row,None) for row in rows)
        return rowgeoms

    def load_reader(self):
        import openpyxl as pyxl

        encoding = self.kwargs.get('encoding', 'utf8')
        wb = pyxl.load_workbook(self.filepath, read_only=True) # WHAT'S THE ENCODING ARG HERE?
        
        if "sheet" in self.kwargs:
            sheet = wb[self.kwargs["sheet"]]
        else:
            sheet = wb[wb.sheetnames[0]]

        # read only has to explicitly calculate dimension
        dimstring = sheet.calculate_dimension()
        lowerright = dimstring.split(':')[-1].strip()
        lastrow = int(float(''.join(c for c in lowerright if c.isdigit())))
        sheet.nrows = lastrow
            
        return sheet

    def load_fields(self):
        rows = ([cell.value for cell in row] for row in self.reader.iter_rows())

        # some excel files may contain junk metadata near top that should be skipped
        # TODO: maybe change API/keywords here...
        if "skip" in self.kwargs:
            for _ in range(self.kwargs["skip"]):
                next(rows)
                
        fieldnames = next(rows)
        return fieldnames

    def load_meta(self):
        return None

    
